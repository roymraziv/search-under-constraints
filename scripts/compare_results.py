#!/usr/bin/env python3
"""
Compare two benchmark result summaries in Excel format.

Usage:
    python scripts/compare_results.py <result_dir1> <result_dir2> [output.xlsx]
"""

import argparse
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError as e:
    print(f"ERROR: Missing required dependency: {e}", file=sys.stderr)
    print("Install with: pip install pandas openpyxl", file=sys.stderr)
    sys.exit(1)


def validate_result_dir(path: Path) -> Path:
    """Validate result directory has summary.csv."""
    if not path.exists():
        raise FileNotFoundError(f"Directory not found: {path}")
    if not path.is_dir():
        raise ValueError(f"Path is not a directory: {path}")
    
    csv_path = path / "summary.csv"
    if not csv_path.exists():
        raise FileNotFoundError(f"summary.csv not found in: {path}")
    
    return csv_path


def main():
    parser = argparse.ArgumentParser(
        description="Compare two benchmark result summaries in Excel format",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/compare_results.py results/2026-02-05_1455 results/2026-02-05_1333
  python scripts/compare_results.py results/run1 results/run2 -o comparison.xlsx
        """
    )
    parser.add_argument("dir1", help="First result directory path")
    parser.add_argument("dir2", help="Second result directory path")
    parser.add_argument(
        "-o", "--output",
        help="Output Excel file path (default: compare_<dir1>_<dir2>.xlsx)",
        default=None
    )
    
    args = parser.parse_args()
    
    try:
        # Validate directories and CSV files
        csv1_path = validate_result_dir(Path(args.dir1))
        csv2_path = validate_result_dir(Path(args.dir2))
        
        # Read CSVs
        print(f"Reading {csv1_path}...")
        df1 = pd.read_csv(csv1_path)
        print(f"Reading {csv2_path}...")
        df2 = pd.read_csv(csv2_path)
        
        # Generate sheet names from directory names (Excel limit: 31 chars)
        sheet1_name = Path(args.dir1).name[:31]
        sheet2_name = Path(args.dir2).name[:31]
        
        # Determine output path
        if args.output:
            output_path = Path(args.output)
        else:
            # Default: compare_<dir1>_<dir2>.xlsx in current directory
            dir1_name = Path(args.dir1).name.replace(" ", "_")
            dir2_name = Path(args.dir2).name.replace(" ", "_")
            output_path = Path(f"compare_{dir1_name}_{dir2_name}.xlsx")
        
        # Write to Excel
        print(f"Creating Excel workbook: {output_path}")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df1.to_excel(writer, sheet_name=sheet1_name, index=False)
            df2.to_excel(writer, sheet_name=sheet2_name, index=False)
        
        # Add formatting
        print("Applying formatting...")
        wb = load_workbook(output_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Freeze header row
            ws.freeze_panes = "A2"
            # Bold headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
        wb.save(output_path)
        
        print(f"✓ Created comparison Excel file: {output_path}")
        print(f"  Sheet 1: {sheet1_name} ({len(df1)} rows)")
        print(f"  Sheet 2: {sheet2_name} ({len(df2)} rows)")
        
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
